"""
Конвертер бизнес-календаря: .xlsx → docs/NN/data/calendar.json

Сквозной принцип: любое расхождение останавливает генерацию с сообщением,
называющим месяц и ячейку. Приложение офлайновое — тихо испорченные данные
никто не заметит месяцами, поэтому падение всегда лучше догадки.

Диапазон, набор строк и дополнительные символы конвертер берёт из конфига
клиента (--client), а не из собственных констант: у клиентов они разные.

Режимы:
  --report-colors   перечисляет все встреченные заливки с их видом и
                    примерами ячеек. Запускается ПЕРВЫМ на новом файле,
                    до того как править палитру в calendar_config.
  (обычный)         пишет docs/NN/data/calendar.json
"""

import argparse
import calendar as pycal
import hashlib
import json
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import date, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries

sys.path.insert(0, str(Path(__file__).resolve().parent))
import calendar_config as cfg
import console
from client_config import ClientConfigError, load_client
from console import fail, head, info, ok, warn
from xlsx_colors import ThemeResolver, describe_fill

SCHEMA_VERSION = 1


class ConvertError(Exception):
    """Остановка с внятной причиной. Текст уходит пользователю как есть."""


# ===========================================================================
# Вспомогательное
# ===========================================================================

def cell_ref(row, col):
    return f"{get_column_letter(col)}{row}"


def norm_text(value):
    """NFC-нормализация плюс схлопывание пробелов и переносов."""
    if value is None:
        return ""
    text = unicodedata.normalize("NFC", str(value))
    return re.sub(r"\s+", " ", text).strip()


def month_key(y, m):
    return f"{y:04d}-{m:02d}"


def day_key(d):
    return d.isoformat()


def iter_months(start, count):
    y, m = start.year, start.month
    for _ in range(count):
        yield y, m
        m += 1
        if m == 13:
            m, y = 1, y + 1


# ===========================================================================
# Выбор листа
# ===========================================================================

def find_calendar_sheet(wb, autodetect=False):
    """Ищем лист с блоками месяцев — по умолчанию строго по имени.

    Автоопределение выключено намеренно. Переименованный лист означает, что
    структура книги поехала, и молча разобрать «похожий» лист — ровно тот
    тихий сбой, ради предотвращения которого написан весь этот конвертер.
    Осознанный обход есть: --allow-sheet-autodetect.
    """
    pattern = re.compile(cfg.MONTH_TITLE_PATTERN)

    def has_blocks(ws):
        found = 0
        for row in ws.iter_rows(min_col=cfg.COL_LABEL, max_col=cfg.COL_LABEL,
                                max_row=min(ws.max_row, 400)):
            text = norm_text(row[0].value)
            if text:
                m = pattern.match(text)
                if m and m.group(1) in cfg.MONTH_NAMES:
                    found += 1
        return found

    for name in cfg.CALENDAR_SHEET_CANDIDATES:
        if name in wb.sheetnames and has_blocks(wb[name]):
            return wb[name]

    matches = [ws for ws in wb.worksheets if has_blocks(ws) >= 2]

    if not autodetect:
        hint = (f"Похоже на календарь: {', '.join(ws.title for ws in matches)}. "
                if matches else "Ни один лист не похож на календарь. ")
        raise ConvertError(
            "не найден лист календаря ни под одним из ожидаемых имён: "
            f"{', '.join(cfg.CALENDAR_SHEET_CANDIDATES)}.\n"
            f"В книге есть листы: {', '.join(wb.sheetnames)}.\n"
            + hint +
            "Переименуйте лист обратно, добавьте его имя в "
            "CALENDAR_SHEET_CANDIDATES или запустите с флагом "
            "--allow-sheet-autodetect, если уверены.")

    if len(matches) == 1:
        warn(f"лист календаря определён автоматически: «{matches[0].title}»")
        return matches[0]
    if not matches:
        raise ConvertError(
            "не найден лист с блоками месяцев. "
            f"В книге есть: {', '.join(wb.sheetnames)}")
    raise ConvertError(
        "блоки месяцев найдены сразу на нескольких листах: "
        f"{', '.join(ws.title for ws in matches)}. "
        "Укажите нужный в CALENDAR_SHEET_CANDIDATES.")


def find_legend_sheet(wb, calendar_ws):
    for name in cfg.LEGEND_SHEET_CANDIDATES:
        if name in wb.sheetnames and wb[name] is not calendar_ws:
            return wb[name]
    for ws in wb.worksheets:
        if ws is not calendar_ws:
            return ws
    return None


# ===========================================================================
# Условное форматирование
# ===========================================================================

def check_conditional_formatting(ws, blocks, client):
    """Останавливаемся, только если правило задевает сами данные.

    Безобидное правило в легенде или служебной части листа блокировать
    работу не должно — проверяем прицельно, по пересечению диапазонов.
    """
    try:
        rules = list(ws.conditional_formatting)
    except Exception:
        return

    if not rules:
        return

    data_rows = set()
    for b in blocks:
        data_rows.add(b["dates_row"])
        for i in range(client.activity_row_count):
            data_rows.add(b["first_activity_row"] + i)

    hits = []
    for cf in rules:
        for rng in str(cf.sqref).split():
            try:
                min_col, min_row, max_col, max_row = range_boundaries(rng)
            except Exception:
                continue
            rows = set(range(min_row, (max_row or min_row) + 1))
            cols_touch = (max_col or min_col) >= cfg.COL_FIRST_DAY
            if rows & data_rows and cols_touch:
                hits.append(rng)

    if hits:
        raise ConvertError(
            "на календарных данных стоит условное форматирование: "
            f"{', '.join(sorted(set(hits)))}.\n"
            "openpyxl видит только обычные заливки, поэтому цвет, созданный "
            "правилом, потеряется и календарь выйдет обесцвеченным.\n"
            "Замените условное форматирование обычной заливкой ячеек.")


# ===========================================================================
# Поиск блоков месяцев
# ===========================================================================

def _looks_like_title_row(ws, row_idx):
    """Название месяца встречается в блоке трижды.

    Кроме собственно шапки оно продублировано в объединённых ячейках строк
    дней недели и дат — так устроен исходный шаблон. Настоящую шапку отличаем
    структурно: ровно под ней, через OFFSET_DATES строк, начинается нумерация
    дней с единицы. У двух самозванцев там окажется строка активностей.
    """
    probe = ws.cell(row=row_idx + cfg.OFFSET_DATES, column=cfg.COL_FIRST_DAY).value
    try:
        return int(str(probe).strip()) == 1
    except (TypeError, ValueError):
        return False


def find_blocks(ws):
    pattern = re.compile(cfg.MONTH_TITLE_PATTERN)
    blocks = []
    rejected = 0

    for row_idx in range(1, ws.max_row + 1):
        text = norm_text(ws.cell(row=row_idx, column=cfg.COL_LABEL).value)
        if not text:
            continue
        m = pattern.match(text)
        if not m:
            continue
        name, year = m.group(1), int(m.group(2))
        if name not in cfg.MONTH_NAMES:
            continue
        if not _looks_like_title_row(ws, row_idx):
            rejected += 1
            continue
        blocks.append({
            "title_row": row_idx,
            "title": text,
            "year": year,
            "month": cfg.MONTH_NAMES[name],
            "note_row": row_idx + cfg.OFFSET_NOTE,
            "dates_row": row_idx + cfg.OFFSET_DATES,
            "first_activity_row": row_idx + cfg.OFFSET_FIRST_ACTIVITY,
        })

    if not blocks:
        raise ConvertError(
            "на листе календаря не найдено ни одного блока месяца.\n"
            f"Заголовков месяцев встречено {rejected}, но ни под одним из них "
            f"через {cfg.OFFSET_DATES} строк нет нумерации дней с единицы — "
            "похоже, раскладка блока изменилась.")
    return blocks


def read_status(ws, block, next_title_row):
    """STATUS: READY в колонке A, между заголовком месяца и следующим.

    Ровно один статус на блок. Опечатка вроде «STATUS: REDY» — ошибка,
    чтобы она не превращалась молча в «месяц не заполнен».
    """
    stop = next_title_row if next_title_row else ws.max_row + 1
    found = []
    for r in range(block["title_row"], stop):
        text = norm_text(ws.cell(row=r, column=cfg.COL_STATUS).value)
        if text:
            found.append((r, text))

    if not found:
        return False
    if len(found) > 1:
        rows = ", ".join(cell_ref(r, cfg.COL_STATUS) for r, _ in found)
        raise ConvertError(
            f"{block['title']}: в колонке A найдено несколько статусов ({rows}). "
            "Допустим ровно один на блок.")

    row, text = found[0]
    if text.upper() != cfg.STATUS_READY.upper():
        raise ConvertError(
            f"{block['title']}: в {cell_ref(row, cfg.COL_STATUS)} стоит «{text}», "
            f"а ожидается «{cfg.STATUS_READY}» либо пустая ячейка.\n"
            "Опечатка в статусе трактовалась бы как «месяц не заполнен», "
            "поэтому конвертер останавливается.")
    return True


def read_dates_row(ws, block):
    """Числа месяца с проверкой непрерывности и длины.

    Ширина блока заранее не предполагается: месяцы бывают по 28, 30 и 31
    дню, а шаблон сделан копированием августа — оставшаяся от него лишняя
    «31» ловится именно здесь.
    """
    row = block["dates_row"]
    title = block["title"]
    found = []

    for col in range(cfg.COL_FIRST_DAY, cfg.COL_FIRST_DAY + cfg.MAX_DAYS + 5):
        value = ws.cell(row=row, column=col).value
        if value is None or (isinstance(value, str) and not value.strip()):
            continue
        try:
            num = int(str(value).strip())
        except (TypeError, ValueError):
            raise ConvertError(
                f"{title}: в строке дат {cell_ref(row, col)} значение «{value}» — не число.")
        found.append((col, num))

    if not found:
        raise ConvertError(f"{title}: строка дат (строка {row}) пуста.")

    numbers = [n for _, n in found]
    expected_len = pycal.monthrange(block["year"], block["month"])[1]

    duplicates = [n for n, c in Counter(numbers).items() if c > 1]
    if duplicates:
        raise ConvertError(
            f"{title}: числа повторяются — {sorted(duplicates)}. "
            "Скорее всего колонка продублирована при копировании.")

    if numbers != list(range(1, len(numbers) + 1)):
        raise ConvertError(
            f"{title}: числа идут не подряд от 1. Получено: {numbers}.\n"
            "Пропуск в середине ряда обрубил бы месяц при наивном чтении, "
            "поэтому проверяется непрерывность.")

    if len(numbers) != expected_len:
        raise ConvertError(
            f"{title}: найдено дней {len(numbers)}, а в месяце их {expected_len}.\n"
            + ("Февраль 2027 не високосный — 28 дней; лишние колонки от "
               "августовского шаблона надо убрать."
               if block["month"] == 2 else
               "Проверьте, не осталась ли лишняя колонка от копии августа."))

    day_to_col = {n: c for c, n in found}
    last_col = day_to_col[expected_len]
    return day_to_col, last_col


def check_nothing_beyond(ws, block, last_col, resolver, client):
    """Справа от последнего дня не должно быть ни отметок, ни заливок.

    Ловит остатки августовской копии: число стёрли, а зелёный фон 31-й
    колонки остался.
    """
    title = block["title"]
    rows = [block["dates_row"]] + [block["first_activity_row"] + i
                                   for i in range(client.activity_row_count)]
    leftovers = []
    for r in rows:
        for col in range(last_col + 1, last_col + 1 + 4):
            cell = ws.cell(row=r, column=col)
            hex6, kind = describe_fill(cell, resolver)
            if norm_text(cell.value) or (hex6 and kind != "none"):
                leftovers.append(cell_ref(r, col))
    if leftovers:
        raise ConvertError(
            f"{title}: справа от последнего дня остались размеченные ячейки — "
            f"{', '.join(leftovers[:8])}"
            f"{' и ещё…' if len(leftovers) > 8 else ''}.\n"
            "Это следы копии августа: очистите значение и заливку.")


# ===========================================================================
# Символы
# ===========================================================================

def parse_marks(raw, title, ref, client):
    """Нормализует содержимое ячейки и разбирает его на известные символы.

    Каталог символов — общий для астролога, но клиент может объявить свои
    буквы в extraMarks: у одного в строках массаж, у другого пластика.
    """
    text = norm_text(raw)
    if not text:
        return ""

    known = client.known_marks
    out = []
    for ch in text:
        if ch == " ":
            continue
        ch = cfg.HOMOGLYPHS.get(ch, ch)
        if ch not in known:
            raise ConvertError(
                f"{title}: в {ref} встретился неизвестный символ «{ch}» "
                f"(код U+{ord(ch):04X}) в значении «{text}».\n"
                f"Известные: {' '.join(known)}.\n"
                "Символ без расшифровки не должен доехать до приложения.\n"
                "Если буква законная, объявите её в extraMarks конфига клиента.")
        out.append(ch)
    return "".join(out)


# ===========================================================================
# Цвета
# ===========================================================================

def bucket_for(hex6, kind, palette, title, ref, zone):
    if hex6 is None:
        return None
    bucket = palette.get(hex6.upper())
    if bucket is None:
        raise ConvertError(
            f"{title}: в {ref} заливка #{hex6} (вид: {kind}) не описана "
            f"в палитре зоны «{zone}».\n"
            "Добавьте оттенок в calendar_config.py или замените его в таблице "
            "на один из уже используемых.\n"
            "Полный список оттенков файла: python tools/xlsx_to_calendar_json.py "
            "--report-colors")
    return bucket


# ===========================================================================
# Разбор блока
# ===========================================================================

def parse_block(ws, block, next_title_row, resolver, client):
    title = block["title"]
    count = client.activity_row_count
    ready = read_status(ws, block, next_title_row)
    day_to_col, last_col = read_dates_row(ws, block)
    check_nothing_beyond(ws, block, last_col, resolver, client)

    note = norm_text(ws.cell(row=block["note_row"], column=cfg.COL_LABEL).value)

    # Число строк объявлено в конфиге клиента. Расхождение — стоп: у одного
    # клиента строк восемь, у другого может быть иначе, и «догадаться»
    # означало бы молча потерять или выдумать строку.
    labels = []
    for i in range(count):
        r = block["first_activity_row"] + i
        labels.append(norm_text(ws.cell(row=r, column=cfg.COL_LABEL).value))
    if not all(labels):
        missing = [i + 1 for i, v in enumerate(labels) if not v]
        raise ConvertError(
            f"{title}: не найдены подписи строк активностей № {missing} "
            f"(строки {block['first_activity_row']}–"
            f"{block['first_activity_row'] + count - 1}).\n"
            f"Конфиг клиента объявляет {count} строк подряд под строкой дат.")

    extra_row = block["first_activity_row"] + count
    if norm_text(ws.cell(row=extra_row, column=cfg.COL_LABEL).value):
        raise ConvertError(
            f"{title}: под объявленными {count} строками активностей нашлась "
            f"ещё одна — {cell_ref(extra_row, cfg.COL_LABEL)}. "
            "Либо раскладка блока изменилась, либо строку забыли добавить "
            "в конфиг клиента; разбор остановлен.")

    # Фазы луны из заливок строки дат
    moon = {}
    for day, col in day_to_col.items():
        cell = ws.cell(row=block["dates_row"], column=col)
        hex6, kind = describe_fill(cell, resolver)
        bucket = bucket_for(hex6, kind, cfg.PALETTE_DATES, title,
                            cell_ref(block["dates_row"], col), "строка дат")
        if bucket == "moon_full":
            moon[day] = "full"
        elif bucket == "moon_new":
            moon[day] = "new"

    # Сетка активностей
    days = {}
    for day, col in day_to_col.items():
        cells = {}
        for i, rowdef in enumerate(client.rows):
            r = block["first_activity_row"] + i
            cell = ws.cell(row=r, column=col)
            ref = cell_ref(r, col)
            hex6, kind = describe_fill(cell, resolver)
            bucket = bucket_for(hex6, kind, cfg.PALETTE_ACTIVITY, title, ref, "активности")
            cells[rowdef["id"]] = {
                "c": bucket or "none",
                "m": parse_marks(cell.value, title, ref, client),
            }
        days[day] = {"cells": cells}

    return {
        "key": month_key(block["year"], block["month"]),
        "year": block["year"],
        "month": block["month"],
        "title": f"{cfg.MONTH_TITLES_RU[block['month']]} {block['year']}",
        "note": note,
        "ready": ready,
        "days": days,
        "moon": moon,
        "labels": labels,
    }


# ===========================================================================
# Заполненность
# ===========================================================================

def resolve_ready(parsed, august_signature):
    """Статус из таблицы — основной признак, сетка — вспомогательный.

    Заметка сама по себе месяц не хоронит: заполненный месяц с пустой
    заметкой остаётся заполненным.
    """
    for m in parsed:
        note_is_placeholder = (
            not m["note"]
            or norm_text(m["note"]).lower() == norm_text(cfg.PLACEHOLDER_NOTE).lower()
        )
        signature = json.dumps(m["days"], sort_keys=True, ensure_ascii=False)
        same_as_august = (m["key"] != "2026-08" and signature == august_signature)

        if m["ready"]:
            if same_as_august:
                warn(f"{m['title']}: статус READY, но сетка совпадает с августовской — "
                     "проверьте, точно ли месяц заполнен")
            if note_is_placeholder:
                warn(f"{m['title']}: статус READY, но заметка пуста или осталась "
                     "плейсхолдером")
        else:
            if not note_is_placeholder and not same_as_august:
                warn(f"{m['title']}: статуса нет, но заметка и сетка выглядят "
                     "заполненными — месяц помечен как незаполненный")


# ===========================================================================
# Фазы луны: чередование и разрывы
# ===========================================================================

def add_moon_anchors(events, client):
    """Приставляет к списку фазы, лежащие сразу за краями диапазона.

    Зачем: интерполяция требует опоры с обеих сторон, а таблица астролога
    заканчивается вместе с диапазоном. Без опор первые и последние две
    недели календаря остаются без фазы — до четверти месяца на каждом краю.

    Опора приходит данными, а не расчётом. Средняя синодическая модель
    промахивается на сутки (в этом проекте она уже промахнулась по августу
    2026), а ошибка в опоре тихо перекашивает освещённость на две недели
    вперёд.

    Опора обязана быть БЛИЖАЙШЕЙ фазой снаружи. Если между ней и краем
    таблицы прячется ещё одна, интерполяция пройдёт сквозь неё и нарисует
    диск задом наперёд — поэтому проверяем и чередование, и расстояние.
    """
    if not events:
        return events

    for side, neighbour in (("before", events[0]), ("after", events[-1])):
        anchor = client.moon_anchors.get(side)
        if anchor is None:
            continue

        gap = abs((date.fromisoformat(anchor["date"])
                   - date.fromisoformat(neighbour["date"])).days)
        if gap > cfg.MOON_GAP_BREAK:
            raise ConvertError(
                f"опора луны moonAnchors.{side} ({anchor['type']} "
                f"{anchor['date']}) отстоит от ближайшей фазы таблицы "
                f"({neighbour['type']} {neighbour['date']}) на {gap} дней, "
                f"а между соседними фазами бывает не больше "
                f"{cfg.MOON_GAP_BREAK}.\n"
                "Похоже, названа не ближайшая фаза: между ней и краем "
                "таблицы есть ещё одна. Нужна именно соседняя.")

        events = [anchor] + events if side == "before" else events + [anchor]

    return events


def build_moon(parsed, client):
    events = []
    for m in parsed:
        for day, kind in sorted(m["moon"].items()):
            events.append({"date": day_key(date(m["year"], m["month"], day)), "type": kind})
    events.sort(key=lambda e: e["date"])

    events = add_moon_anchors(events, client)

    for a, b in zip(events, events[1:]):
        if a["type"] == b["type"]:
            raise ConvertError(
                f"две одинаковые фазы подряд: {a['type']} {a['date']} и "
                f"{b['type']} {b['date']}.\n"
                "Новолуния и полнолуния обязаны чередоваться; скорее всего "
                "перепутан оттенок синего в строке дат.\n"
                "Без чередования направление фазы считается неверно.")

    gaps = []
    for a, b in zip(events, events[1:]):
        d1 = date.fromisoformat(a["date"])
        d2 = date.fromisoformat(b["date"])
        delta = (d2 - d1).days
        if delta > cfg.MOON_GAP_BREAK:
            gaps.append({"from": a["date"], "to": b["date"]})
            warn(f"разрыв в фазах: {delta} дней между {a['date']} и {b['date']} — "
                 "на этом отрезке фаза показываться не будет")
        elif cfg.MOON_GAP_WARN[0] <= delta <= cfg.MOON_GAP_WARN[1]:
            warn(f"необычный промежуток фаз: {delta} дней между "
                 f"{a['date']} и {b['date']}")

    if events:
        low = cfg.MOON_EVENTS_EXPECTED - cfg.MOON_EVENTS_TOLERANCE
        high = cfg.MOON_EVENTS_EXPECTED + cfg.MOON_EVENTS_TOLERANCE
        if not (low <= len(events) <= high):
            warn(f"фаз луны найдено {len(events)}, а за год ожидается около "
                 f"{cfg.MOON_EVENTS_EXPECTED}")
    else:
        warn("фазы луны в таблице не размечены вовсе")

    return events, gaps


# ===========================================================================
# Легенда
# ===========================================================================

def parse_legend(ws, client):
    """Тексты расшифровок из таблицы. Набор символов сюда не заглядывает —
    он задан каталогом и конфигом клиента, иначе сдвиг легендного блока
    сделал бы неизвестными разом все символы."""
    texts = {}
    if ws is None:
        return texts

    known = client.known_marks
    pattern = re.compile(r"^(.)\s*[–—-]\s*(.+)$")
    for row in ws.iter_rows():
        for cell in row:
            line = norm_text(cell.value)
            if len(line) < 4:
                continue
            m = pattern.match(line)
            if not m:
                continue
            ch = cfg.HOMOGLYPHS.get(m.group(1), m.group(1))
            if ch in known and ch not in texts:
                texts[ch] = m.group(2).strip()
    return texts


def build_legend(legend_texts, used_marks, client):
    fallback = client.mark_texts
    marks = []
    for ch in client.known_marks:
        text = legend_texts.get(ch) or fallback[ch]
        if ch in used_marks and ch not in legend_texts:
            warn(f"символ «{ch}» есть в сетке, но описания в легенде таблицы "
                 "не нашлось — взят резервный текст из конфига")
        if ch not in used_marks:
            info(f"символ «{ch}» описан в легенде, но ни разу не встретился в сетке")
        marks.append({"char": ch, "text": text})

    return {
        "colors": [{"key": k, "text": v} for k, v in cfg.COLOR_TEXTS.items()],
        "marks": marks,
        "moon": [{"key": k, "text": v} for k, v in cfg.MOON_TEXTS.items()],
    }


# ===========================================================================
# Проверка диапазона
# ===========================================================================

def validate_range(parsed, days, client):
    start = client.start_date
    end = client.end_date

    keys = [m["key"] for m in parsed]
    expected = [month_key(y, m) for y, m in iter_months(start, client.expected_months)]

    dupes = [k for k, c in Counter(keys).items() if c > 1]
    if dupes:
        raise ConvertError(
            f"месяцы повторяются: {sorted(dupes)}. "
            "Скорее всего блок скопирован, а заголовок не поправлен.")

    # Сверку набора делаем раньше подсчёта: так сообщение называет
    # конкретный месяц, а не только расхождение в количестве.
    if keys != expected:
        missing = [k for k in expected if k not in keys]
        extra = [k for k in keys if k not in expected]
        parts = []
        if missing:
            parts.append(f"не хватает: {', '.join(missing)}")
        if extra:
            parts.append(f"лишние: {', '.join(extra)}")
        if not parts:
            parts.append(f"нарушен порядок: {', '.join(keys)}")
        raise ConvertError(
            f"набор месяцев не совпадает с объявленным в конфиге клиента "
            f"({client.range_start[:7]} … {client.range_end[:7]}). "
            + "; ".join(parts))

    if len(parsed) != client.expected_months:
        raise ConvertError(
            f"найдено блоков месяцев: {len(parsed)}, а конфиг клиента "
            f"объявляет {client.expected_months}.")

    if len(days) != client.expected_days:
        raise ConvertError(
            f"дней в календаре {len(days)}, а конфиг клиента объявляет "
            f"{client.expected_days}.")

    cur = start
    while cur <= end:
        if day_key(cur) not in days:
            raise ConvertError(f"в диапазоне пропущен день {day_key(cur)}")
        cur += timedelta(days=1)

    for k in days:
        d = date.fromisoformat(k)
        if not (start <= d <= end):
            raise ConvertError(f"день {k} выходит за диапазон "
                               f"{client.range_start}…{client.range_end}")


# ===========================================================================
# Хеш данных
# ===========================================================================

def compute_hash(payload):
    """Каноническая сериализация с фиксированным порядком ключей.

    Из хеша исключены dataHash и generatedAt: иначе получилась бы круговая
    зависимость, а холостой перегон менял бы версию данных на ровном месте.
    """
    subset = {k: v for k, v in payload.items() if k not in ("dataHash", "generatedAt")}
    blob = json.dumps(subset, sort_keys=True, ensure_ascii=False,
                      separators=(",", ":")).encode("utf-8")
    return "sha256:" + hashlib.sha256(blob).hexdigest()


# ===========================================================================
# Приватность
# ===========================================================================

def scrub(text):
    out = text
    for word in cfg.PRIVACY_STOPWORDS:
        out = re.sub(re.escape(word), "", out, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", out).replace(" ,", ",").replace(" .", ".").strip()


# ===========================================================================
# Режим --report-colors
# ===========================================================================

def report_colors(ws, blocks, resolver, client):
    head("Заливки, встреченные в файле")

    zones = {"строка дат": defaultdict(list), "активности": defaultdict(list)}
    kinds = Counter()

    for block in blocks:
        try:
            day_to_col, _ = read_dates_row(ws, block)
        except ConvertError as exc:
            warn(f"{block['title']}: строку дат разобрать не удалось ({exc}); "
                 "блок пропущен в отчёте")
            continue

        for _, col in day_to_col.items():
            targets = [(block["dates_row"], "строка дат")] + [
                (block["first_activity_row"] + i, "активности")
                for i in range(client.activity_row_count)]
            for r, zone in targets:
                hex6, kind = describe_fill(ws.cell(row=r, column=col), resolver)
                kinds[kind] += 1
                if hex6:
                    zones[zone][hex6].append(cell_ref(r, col))

    for kind, count in kinds.most_common():
        info(f"вид заливки «{kind}»: {count} ячеек")

    for zone, palette in (("строка дат", cfg.PALETTE_DATES),
                          ("активности", cfg.PALETTE_ACTIVITY)):
        head(f"Зона: {zone}")
        table = zones[zone]
        if not table:
            info("заливок не найдено")
            continue
        for hex6, refs in sorted(table.items(), key=lambda kv: -len(kv[1])):
            known = palette.get(hex6.upper())
            mark = f"→ {known}" if known else "→ НЕ ОПИСАН В ПАЛИТРЕ"
            sample = ", ".join(refs[:4])
            info(f"#{hex6}  {len(refs):>5} ячеек  {mark}   примеры: {sample}")

    head("Что дальше")
    if cfg.PROVISIONAL_PALETTE:
        warn("PROVISIONAL_PALETTE = True — палитра ещё не подтверждена по реальному файлу")
        info("Сверьте оттенки выше с таблицей, поправьте PALETTE_* в calendar_config.py")
        info("и переведите PROVISIONAL_PALETTE в False.")
    else:
        ok("палитра помечена как подтверждённая")


# ===========================================================================
# Основной проход
# ===========================================================================

def convert(xlsx_path, out_path, client, report_only=False, autodetect=False):
    wb = load_workbook(xlsx_path, data_only=True)
    resolver = ThemeResolver(wb)

    ws = find_calendar_sheet(wb, autodetect=autodetect)
    info(f"лист календаря: «{ws.title}»")
    blocks = find_blocks(ws)
    info(f"найдено блоков месяцев: {len(blocks)}")

    if report_only:
        report_colors(ws, blocks, resolver, client)
        return None

    check_conditional_formatting(ws, blocks, client)

    parsed = []
    for i, block in enumerate(blocks):
        nxt = blocks[i + 1]["title_row"] if i + 1 < len(blocks) else None
        parsed.append(parse_block(ws, block, nxt, resolver, client))

    august = next((m for m in parsed if m["key"] == "2026-08"), None)
    august_signature = (json.dumps(august["days"], sort_keys=True, ensure_ascii=False)
                        if august else None)
    resolve_ready(parsed, august_signature)

    days = {}
    used_marks = set()
    for m in parsed:
        for day, payload in m["days"].items():
            key = day_key(date(m["year"], m["month"], day))
            days[key] = payload
            for cell in payload["cells"].values():
                used_marks.update(cell["m"])

    validate_range(parsed, days, client)

    moon_events, moon_gaps = build_moon(parsed, client)
    legend = build_legend(parse_legend(find_legend_sheet(wb, ws), client),
                          used_marks, client)

    payload = {
        "version": SCHEMA_VERSION,
        "generatedAt": date.today().isoformat(),
        "dataHash": "",
        "provisional": bool(cfg.PROVISIONAL_PALETTE),
        "range": {"start": client.range_start, "end": client.range_end},
        "rows": [dict(r) for r in client.rows],
        "legend": legend,
        "months": [{"key": m["key"], "title": m["title"],
                    "note": scrub(m["note"]) if m["ready"] else "",
                    "ready": m["ready"]} for m in parsed],
        "moonEvents": moon_events,
        "moonGaps": moon_gaps,
        "days": dict(sorted(days.items())),
    }
    payload["dataHash"] = compute_hash(payload)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=1, sort_keys=False),
        encoding="utf-8")

    return payload


def print_summary(payload, out_path, client):
    head("Результат")
    ok(f"записан {out_path}")
    info(f"дней: {len(payload['days'])}")
    info(f"месяцев: {len(payload['months'])}")
    info(f"фаз луны: {len(payload['moonEvents'])}, разрывов: {len(payload['moonGaps'])}")
    info(f"dataHash: {payload['dataHash'][:23]}…")

    ready = [m["key"] for m in payload["months"] if m["ready"]]
    not_ready = [m["key"] for m in payload["months"] if not m["ready"]]

    head("Заполненность месяцев")
    ok(f"готовы ({len(ready)}): {', '.join(ready) if ready else '—'}")
    if not_ready:
        warn(f"не заполнены ({len(not_ready)}): {', '.join(not_ready)}")

    head("Шлюз перед публикацией")
    gate = [
        (len(not_ready) == 0, "все месяцы имеют статус READY"),
        (len(payload["days"]) == client.expected_days,
         f"ровно {client.expected_days} дней"),
        (len(payload["moonGaps"]) == 0, "разрывов в фазах нет"),
        (not payload["provisional"], "палитра подтверждена по реальному .xlsx"),
        (bool(payload["dataHash"]), "dataHash рассчитан"),
    ]
    for passed, text in gate:
        (ok if passed else fail)(text)

    if all(p for p, _ in gate):
        head("Готово к публикации как финальная версия")
    else:
        head("Это демонстрационная сборка, не финальный календарь")


def main():
    ap = argparse.ArgumentParser(description="xlsx → calendar.json")
    root = Path(__file__).resolve().parents[1]
    ap.add_argument("--xlsx", default=str(root / ".tmp" / "calendar.xlsx"))
    ap.add_argument("--out", default=str(root / "docs" / "01" / "data" / "calendar.json"))
    ap.add_argument("--client", default=str(root / "clients" / "01" / "client.json"),
                    help="конфиг клиента: диапазон, строки, доп. символы")
    ap.add_argument("--report-colors", action="store_true",
                    help="показать все заливки файла и остановиться")
    ap.add_argument("--allow-sheet-autodetect", action="store_true",
                    help="разрешить определить лист календаря автоматически, "
                         "если он переименован")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        fail(f"файл не найден: {xlsx_path}")
        info("Скачайте таблицу: File → Download → Microsoft Excel (.xlsx)")
        return 2

    try:
        client = load_client(Path(args.client))
    except ClientConfigError as exc:
        fail(str(exc))
        return 2

    try:
        payload = convert(xlsx_path, Path(args.out), client,
                          report_only=args.report_colors,
                          autodetect=args.allow_sheet_autodetect)
    except ConvertError as exc:
        head("Конвертация остановлена")
        fail(str(exc))
        return 1

    if payload is not None:
        print_summary(payload, Path(args.out), client)
    return 0


if __name__ == "__main__":
    sys.exit(main())

"""
Генератор эталонной книги бизнес-календаря.

Зачем нужен: настоящий .xlsx от астролога появится позже, а конвертер и
приложение надо писать и проверять уже сейчас. Скрипт строит структурно
точную копию шаблона, поэтому конвертер разрабатывается против той же
раскладки, что будет в реальном файле.

Два режима:

  --template  верно шаблону: заполнен только август 2026, остальные
              двенадцать блоков — пустые сетки с плейсхолдером в заметке.
              Именно так выглядит таблица прямо сейчас.

  --demo      все тринадцать месяцев заполнены синтетическими, но
              правдоподобными данными. Нужен, чтобы отлаживать приложение
              на полном годе: 396 дней, полный набор фаз луны.

Августовские буквы (О, ОХ, КУ, Ф+, Р+, у, В, Т) взяты из настоящей таблицы
один в один. Цвета — провизорные, см. PROVISIONAL_PALETTE в calendar_config.
"""

import argparse
import calendar
import math
import random
import sys
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
import calendar_config as cfg
import console  # noqa: F401  — переключает stdout в UTF-8


# ---------------------------------------------------------------------------
# Настоящие данные августа 2026
# ---------------------------------------------------------------------------

AUGUST_MARKS = {
    "negotiations": {3: "О", 4: "О", 9: "О", 10: "ОХ", 14: "Х", 17: "Х", 21: "О", 22: "О"},
    "finance":      {3: "У", 4: "У", 12: "У", 13: "У", 21: "КУ", 22: "КУ"},
    "audit":        {},
    "launch":       {},
    "haircut":      {13: "Ф", 14: "Ф", 15: "Ф+", 16: "Р+", 17: "Р+", 18: "Г+", 19: "Г",
                     20: "Р", 21: "Р+", 22: "Р", 23: "Ф", 24: "Ф", 25: "Р", 26: "Р+", 27: "Г"},
    "teeth":        {5: "у", 6: "у", 9: "у"},
    "medical":      {},
    "massage":      {d: ("Т" if 13 <= d <= 27 else "В") for d in range(1, 32)},
}

# Цвета августа диапазонами [от, до, корзина]. Сняты с исходного изображения,
# помечены как провизорные и будут заменены после сверки с реальным .xlsx.
AUGUST_COLORS = {
    "negotiations": [(1, 2, "neutral"), (3, 13, "good"), (14, 14, "neutral"),
                     (15, 16, "bad"), (17, 17, "neutral"), (18, 23, "good"), (24, 31, "bad")],
    "finance":      [(1, 2, "neutral"), (3, 13, "good"), (14, 14, "neutral"),
                     (15, 16, "bad"), (17, 23, "good"), (24, 31, "bad")],
    "audit":        [(1, 2, "neutral"), (3, 11, "good"), (12, 16, "neutral"),
                     (17, 23, "good"), (24, 31, "neutral")],
    "launch":       [(1, 8, "neutral"), (9, 16, "good"), (17, 20, "neutral"),
                     (21, 26, "good"), (27, 28, "bad"), (29, 31, "neutral")],
    "haircut":      [(1, 12, "neutral"), (13, 27, "good"), (28, 31, "neutral")],
    "teeth":        [(1, 4, "neutral"), (5, 6, "good"), (7, 8, "neutral"), (9, 9, "good"),
                     (10, 10, "neutral"), (11, 16, "bad"), (17, 31, "neutral")],
    "medical":      [(1, 16, "neutral"), (17, 23, "good"), (24, 28, "bad"), (29, 31, "neutral")],
    "massage":      [(1, 12, "neutral"), (13, 27, "good"), (28, 31, "neutral")],
}

AUGUST_NOTE = (
    "В августе сезон затмений (с 12 по 28 августа). НЕ инициировать события. "
    "Все должно идти своим чередом. НЕ сопротивляться событиям, которые пришли естественно. "
    "Принимать их осознанно. Они важны для будущего. Вблизи 12 августа – Новолуние и затмение "
    "Солнца. Обратить внимание на профессиональные дела, неудовлетворенность может всплыть. "
    "28 августа – Полнолуние и затмение Луны. Конец месяца поднимет наболевшие вопросы личного "
    "характера, семейного. Конец месяца внесет ясность в отношения и в дела, проблемы которых "
    "тянутся с середины мая. Что-то стабилизируется, будут сделаны выводы. Останется то, что "
    "жизнеспособно для будущего. Начнется новая фаза развития. В период затмений возможна резкая "
    "смена правил игры со стороны ключевых партнеров, крупных инвесторов или госорганов. "
    "Не кипятиться, могут спровоцировать!"
)

WEEKDAY_SHORT = ["пн", "вт", "ср", "чт", "пт", "сб", "вс"]


# ---------------------------------------------------------------------------
# Фазы луны
# ---------------------------------------------------------------------------

SYNODIC = 29.530588853
# Опорное новолуние: 6 января 2000, 18:14 UTC — классическая эпоха.
NEW_MOON_EPOCH = 2451550.26

# Август 2026 размечен в настоящей таблице, и её разметка главнее расчёта.
# Средняя синодическая модель промахивается на несколько часов и в этом
# случае перебрасывает новолуние на 13-е, тогда как реальное — 12 августа
# (то самое полное солнечное затмение), а полнолуние 28-го.
MOON_OVERRIDES = {
    date(2026, 8, 12): "new",
    date(2026, 8, 28): "full",
}


def _to_jd(d):
    """Юлианская дата для полудня указанной календарной даты."""
    a = (14 - d.month) // 12
    y = d.year + 4800 - a
    m = d.month + 12 * a - 3
    jdn = (d.day + (153 * m + 2) // 5 + 365 * y + y // 4 - y // 100 + y // 400 - 32045)
    return jdn - 0.5


def _from_jd(jd):
    z = int(jd + 0.5)
    a = z
    if z >= 2299161:
        alpha = int((z - 1867216.25) / 36524.25)
        a = z + 1 + alpha - alpha // 4
    b = a + 1524
    c = int((b - 122.1) / 365.25)
    dd = int(365.25 * c)
    e = int((b - dd) / 30.6001)
    day = b - dd - int(30.6001 * e)
    month = e - 1 if e < 14 else e - 13
    year = c - 4716 if month > 2 else c - 4715
    return date(year, month, day)


def moon_events_in_range(start, end):
    """Новолуния и полнолуния в диапазоне, по средней синодической модели.

    Точность порядка суток — для демонстрационных данных этого достаточно,
    а в реальном файле фазы всё равно берутся из разметки астролога.
    """
    events = []
    k0 = math.floor((_to_jd(start) - NEW_MOON_EPOCH) / SYNODIC) - 2
    for i in range(k0, k0 + 40):
        new_jd = NEW_MOON_EPOCH + i * SYNODIC
        full_jd = new_jd + SYNODIC / 2
        for jd, kind in ((new_jd, "new"), (full_jd, "full")):
            d = _from_jd(jd)
            if start <= d <= end:
                events.append((d, kind))
    events.sort()
    return events


# ---------------------------------------------------------------------------
# Стили
# ---------------------------------------------------------------------------

def _fill(hex6):
    return PatternFill(start_color="FF" + hex6, end_color="FF" + hex6, fill_type="solid")


def _palette_hex(bucket, palette):
    for hex6, name in palette.items():
        if name == bucket:
            return hex6
    raise KeyError(f"в палитре нет корзины {bucket}")


FILL_GOOD = _fill(_palette_hex("good", cfg.PALETTE_ACTIVITY))
FILL_NEUTRAL = _fill(_palette_hex("neutral", cfg.PALETTE_ACTIVITY))
FILL_BAD = _fill(_palette_hex("bad", cfg.PALETTE_ACTIVITY))
FILL_WEEKEND = _fill(_palette_hex("weekend", cfg.PALETTE_DATES))
FILL_MOON_FULL = _fill(_palette_hex("moon_full", cfg.PALETTE_DATES))
FILL_MOON_NEW = _fill(_palette_hex("moon_new", cfg.PALETTE_DATES))

BUCKET_FILLS = {"good": FILL_GOOD, "neutral": FILL_NEUTRAL, "bad": FILL_BAD}

THIN = Side(style="thin", color="FF999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")


# ---------------------------------------------------------------------------
# Генерация синтетического месяца
# ---------------------------------------------------------------------------

def synth_month(year, month, rng):
    """Правдоподобные, но выдуманные данные для демо-режима."""
    n = calendar.monthrange(year, month)[1]
    colors, marks = {}, {}

    row_marks = {
        "negotiations": ["О", "Х", "ОХ"],
        "finance": ["У", "К", "КУ"],
        "audit": [],
        "launch": [],
        "haircut": ["Ф", "Р", "Г", "Ф+", "Р+", "Г+"],
        "teeth": ["у"],
        "medical": [],
        "massage": ["В", "Т"],
    }

    for row in cfg.ROW_DEFS:
        rid = row["id"]
        col, mrk = {}, {}
        day = 1
        while day <= n:
            bucket = rng.choices(["good", "neutral", "bad"], weights=[4, 5, 2])[0]
            span = rng.randint(2, 6)
            for d in range(day, min(day + span, n + 1)):
                col[d] = bucket
            day += span
        if rid == "massage":
            for d in range(1, n + 1):
                mrk[d] = "Т" if d % 29 < 15 else "В"
        elif row_marks[rid]:
            for d in range(1, n + 1):
                if rng.random() < 0.22:
                    mrk[d] = rng.choice(row_marks[rid])
        colors[rid] = col
        marks[rid] = mrk
    return colors, marks


def august_colors_by_day():
    out = {}
    for rid, ranges in AUGUST_COLORS.items():
        per_day = {}
        for lo, hi, bucket in ranges:
            for d in range(lo, hi + 1):
                per_day[d] = bucket
        out[rid] = per_day
    return out


# ---------------------------------------------------------------------------
# Запись блока месяца
# ---------------------------------------------------------------------------

def write_month_block(ws, top, year, month, colors, marks, note, ready, moon_by_day):
    n = calendar.monthrange(year, month)[1]
    last_col = cfg.COL_FIRST_DAY + n - 1

    title = f"{[k for k, v in cfg.MONTH_NAMES.items() if v == month][0]} {year}"
    ws.cell(row=top, column=cfg.COL_LABEL, value=title).font = Font(bold=True, size=13)

    if ready:
        ws.cell(row=top, column=cfg.COL_STATUS, value=cfg.STATUS_READY)

    note_row = top + cfg.OFFSET_NOTE
    ws.cell(row=note_row, column=cfg.COL_LABEL, value=note)
    ws.merge_cells(start_row=note_row, start_column=cfg.COL_LABEL,
                   end_row=note_row, end_column=max(last_col, cfg.COL_LABEL + 1))

    wd_row = top + cfg.OFFSET_WEEKDAYS
    dt_row = top + cfg.OFFSET_DATES
    ws.cell(row=wd_row, column=cfg.COL_LABEL, value=title)
    ws.cell(row=dt_row, column=cfg.COL_LABEL, value=title)

    for d in range(1, n + 1):
        col = cfg.COL_FIRST_DAY + d - 1
        cur = date(year, month, d)
        weekday = cur.weekday()

        c_wd = ws.cell(row=wd_row, column=col, value=WEEKDAY_SHORT[weekday])
        c_wd.alignment, c_wd.border = CENTER, BORDER

        c_dt = ws.cell(row=dt_row, column=col, value=d)
        c_dt.alignment, c_dt.border = CENTER, BORDER

        phase = moon_by_day.get(cur)
        if phase == "full":
            c_dt.fill = FILL_MOON_FULL
            c_dt.font = Font(color="FFFFFFFF", bold=True)
            c_wd.fill = FILL_MOON_FULL
        elif phase == "new":
            c_dt.fill = FILL_MOON_NEW
            c_wd.fill = FILL_MOON_NEW
        elif weekday >= 5:
            c_dt.fill = FILL_WEEKEND
            c_wd.fill = FILL_WEEKEND

    for i, row in enumerate(cfg.ROW_DEFS):
        r = top + cfg.OFFSET_FIRST_ACTIVITY + i
        ws.cell(row=r, column=cfg.COL_LABEL, value=row["label"])
        for d in range(1, n + 1):
            col = cfg.COL_FIRST_DAY + d - 1
            cell = ws.cell(row=r, column=col)
            bucket = colors.get(row["id"], {}).get(d)
            if bucket:
                cell.fill = BUCKET_FILLS[bucket]
            mark = marks.get(row["id"], {}).get(d)
            if mark:
                cell.value = mark
            cell.alignment, cell.border = CENTER, BORDER

    return top + cfg.OFFSET_FIRST_ACTIVITY + cfg.ACTIVITY_ROW_COUNT + 2


# ---------------------------------------------------------------------------
# Легенда
# ---------------------------------------------------------------------------

def write_legend(ws):
    ws.cell(row=2, column=2,
            value='Легенда таблицы "Бизнес-календарь на год" (Август 2026 - Август 2027)')
    ws.cell(row=4, column=2, value="Пояснения к таблице:")
    ws.cell(row=5, column=2, value="Синий – Полнолуние")
    ws.cell(row=6, column=2, value="Голубой - Новолуние")

    r = 8
    for ch in cfg.KNOWN_MARKS:
        ws.cell(row=r, column=2, value=f"{ch} – {cfg.MARK_TEXTS_FALLBACK[ch]}")
        r += 1
    ws.column_dimensions["B"].width = 90


# ---------------------------------------------------------------------------
# Сборка книги
# ---------------------------------------------------------------------------

def build(mode, out_path):
    start = date.fromisoformat(cfg.RANGE_START)
    end = date.fromisoformat(cfg.RANGE_END)

    events = moon_events_in_range(start, end)
    moon_by_day = {d: kind for d, kind in events}

    # Разметка августа берётся из настоящей таблицы, а не из расчёта.
    moon_by_day = {d: k for d, k in moon_by_day.items()
                   if not (d.year == 2026 and d.month == 8)}
    moon_by_day.update(MOON_OVERRIDES)

    # В шаблонном режиме размечен только август — как в реальной таблице.
    if mode == "template":
        moon_by_day = {d: k for d, k in moon_by_day.items()
                       if d.year == 2026 and d.month == 8}

    wb = Workbook()
    legend = wb.active
    legend.title = cfg.LEGEND_SHEET_CANDIDATES[0]
    write_legend(legend)

    ws = wb.create_sheet(cfg.CALENDAR_SHEET_CANDIDATES[0])
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 30
    for i in range(cfg.COL_FIRST_DAY, cfg.COL_FIRST_DAY + cfg.MAX_DAYS):
        ws.column_dimensions[get_column_letter(i)].width = 4.5

    rng = random.Random(20260801)
    aug_colors = august_colors_by_day()

    top = 2
    y, m = start.year, start.month
    for _ in range(cfg.EXPECTED_MONTHS):
        is_august_2026 = (y == 2026 and m == 8)

        if is_august_2026:
            colors, marks, note, ready = aug_colors, AUGUST_MARKS, AUGUST_NOTE, True
        elif mode == "demo":
            colors, marks = synth_month(y, m, rng)
            note = (f"Демонстрационный прогноз на "
                    f"{cfg.MONTH_TITLES_RU[m].lower()} {y}. Синтетический текст, "
                    f"нужен только чтобы проверить вёрстку и длинные абзацы.")
            ready = True
        else:
            colors, marks, note, ready = {}, {}, cfg.PLACEHOLDER_NOTE, False

        top = write_month_block(ws, top, y, m, colors, marks, note, ready, moon_by_day)
        m += 1
        if m == 13:
            m, y = 1, y + 1

    wb.save(out_path)
    return len(events), moon_by_day


def main():
    ap = argparse.ArgumentParser(description="Генератор эталонной книги календаря")
    ap.add_argument("--mode", choices=["template", "demo"], default="template")
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    out = Path(args.out) if args.out else (
        Path(__file__).resolve().parents[1] / ".tmp" / f"reference-{args.mode}.xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)

    total_events, moon_by_day = build(args.mode, out)

    print(f"Записано: {out}")
    print(f"Режим: {args.mode}")
    print(f"Фаз луны в диапазоне (расчётных): {total_events}")
    print(f"Размечено в книге: {len(moon_by_day)}")
    aug = sorted(d for d in moon_by_day if d.year == 2026 and d.month == 8)
    print(f"Август 2026: {', '.join(f'{d.day} — {moon_by_day[d]}' for d in aug)}")


if __name__ == "__main__":
    main()

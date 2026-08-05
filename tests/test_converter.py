"""
Проверка конвертера на намеренно сломанных книгах.

Каждый тест берёт эталонную книгу, вносит ровно одну поломку и требует,
чтобы конвертер остановился с внятным сообщением. Смысл не в том, что он
упадёт, а в том, что упадёт по нужной причине и назовёт месяц или ячейку:
молча испорченные данные в офлайновом приложении не заметит никто.

Запуск:  python tests/test_converter.py
"""

import shutil
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "tools"))

import calendar_config as cfg
import console
from client_config import ClientConfig, load_client
from console import BOLD, DIM, RESET, fail, head, info, ok
from make_reference_xlsx import build
from xlsx_to_calendar_json import ConvertError, convert, resolve_paths

WORK = ROOT / ".tmp" / "fixtures"
BASE = WORK / "base-demo.xlsx"
# Та же книга, но в раскладке настоящего файла астролога: вкладка на месяц.
TABS = WORK / "base-tabs.xlsx"

CLIENT = load_client(ROOT / "tests" / "fixtures" / "client-test.json")
ROWS = CLIENT.rows


def client_with(**over):
    """Конфиг того же клиента с одним изменённым полем.

    Нужен кейсам, где ломается не книга, а конфиг: конвертер обязан
    остановиться и на расхождении «таблица против конфига», иначе разница
    доехала бы до приложения молча.
    """
    data = {
        "range": {"start": CLIENT.range_start, "end": CLIENT.range_end},
        "rows": [dict(r) for r in ROWS],
        "extraMarks": dict(CLIENT.extra_marks),
    }
    data.update(over)
    return ClientConfig.from_dict(data)


# ---------------------------------------------------------------------------
# Инструменты поломки
# ---------------------------------------------------------------------------

def fresh_copy(name, base=BASE):
    dst = WORK / f"{name}.xlsx"
    shutil.copyfile(base, dst)
    return dst


def calendar_ws(wb):
    return wb[cfg.CALENDAR_SHEET_CANDIDATES[0]]


def tabs_ws(wb, title):
    """Вкладка месяца в книге с раскладкой «вкладка на месяц».

    Ищем по заголовку в ячейке, а не по имени вкладки: в настоящем файле
    астролога вкладка называется «Февраля 2027», а в B2 стоит правильное
    «ФЕВРАЛЬ 2027». Месяц определяется содержимым, имя вкладки — свободное.
    """
    for ws in wb.worksheets:
        cell = ws.cell(row=cfg.TABS_TITLE_ROW, column=cfg.COL_LABEL).value
        if str(cell or "").strip() == title:
            return ws
    raise AssertionError(f"не найдена вкладка {title}")


def find_title_row(ws, title):
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(row=r, column=cfg.COL_LABEL).value or "").strip() == title:
            probe = ws.cell(row=r + cfg.OFFSET_DATES, column=cfg.COL_FIRST_DAY).value
            try:
                if int(str(probe).strip()) == 1:
                    return r
            except (TypeError, ValueError):
                continue
    raise AssertionError(f"не найден блок {title}")


def edit(name, mutate, base=BASE):
    """Копирует эталон, применяет поломку, возвращает путь.

    Книге со стопкой блоков поломка получает готовый лист календаря — их
    девятнадцать, и каждая правит один и тот же лист. Книге со вкладками
    лист не передаётся: она сама выбирает нужную вкладку по месяцу.
    """
    path = fresh_copy(name, base)
    wb = load_workbook(path)
    mutate(wb, calendar_ws(wb) if base == BASE else None)
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Поломки
# ---------------------------------------------------------------------------

def break_missing_day(wb, ws):
    r = find_title_row(ws, "СЕНТЯБРЬ 2026") + cfg.OFFSET_DATES
    ws.cell(row=r, column=cfg.COL_FIRST_DAY + 14).value = None


def break_duplicate_day(wb, ws):
    r = find_title_row(ws, "ОКТЯБРЬ 2026") + cfg.OFFSET_DATES
    ws.cell(row=r, column=cfg.COL_FIRST_DAY + 9).value = 9


def break_february_length(wb, ws):
    top = find_title_row(ws, "ФЕВРАЛЬ 2027")
    r = top + cfg.OFFSET_DATES
    # Дорисовываем 29-й день, которого в 2027 году нет.
    ws.cell(row=r, column=cfg.COL_FIRST_DAY + 28).value = 29


def break_missing_month(wb, ws):
    """Блок марта перестаёт опознаваться целиком.

    Статус в колонке A тоже убираем: иначе он попадёт в диапазон поиска
    предыдущего месяца и первой сработает проверка «несколько статусов»,
    а мы хотим проверить именно контроль набора месяцев.
    """
    top = find_title_row(ws, "МАРТ 2027")
    ws.cell(row=top, column=cfg.COL_LABEL).value = None
    ws.cell(row=top, column=cfg.COL_STATUS).value = None
    ws.cell(row=top + cfg.OFFSET_DATES, column=cfg.COL_FIRST_DAY).value = None


def break_duplicate_month(wb, ws):
    """Два блока с одним ключом месяца.

    Берём май и март — оба по 31 дню. Если подменить месяц на другой по
    длине, первой сработает проверка числа дней, и до контроля дублей
    дело не дойдёт.
    """
    top = find_title_row(ws, "МАЙ 2027")
    ws.cell(row=top, column=cfg.COL_LABEL).value = "МАРТ 2027"


def break_unknown_color(wb, ws):
    top = find_title_row(ws, "МАЙ 2027")
    cell = ws.cell(row=top + cfg.OFFSET_FIRST_ACTIVITY, column=cfg.COL_FIRST_DAY + 3)
    cell.fill = PatternFill(start_color="FF7B1FA2", end_color="FF7B1FA2", fill_type="solid")


def break_conditional_formatting(wb, ws):
    top = find_title_row(ws, "ИЮНЬ 2027")
    r1 = top + cfg.OFFSET_FIRST_ACTIVITY
    r2 = r1 + CLIENT.activity_row_count - 1
    rng = f"C{r1}:AG{r2}"
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="equal", formula=['"О"'],
                   fill=PatternFill(start_color="FF00FF00", end_color="FF00FF00",
                                    fill_type="solid")))


def break_unknown_mark(wb, ws):
    top = find_title_row(ws, "ИЮЛЬ 2027")
    ws.cell(row=top + cfg.OFFSET_FIRST_ACTIVITY, column=cfg.COL_FIRST_DAY + 5).value = "Ж"


def break_status_typo(wb, ws):
    top = find_title_row(ws, "ДЕКАБРЬ 2026")
    ws.cell(row=top, column=cfg.COL_STATUS).value = "STATUS: REDY"


def break_double_status(wb, ws):
    top = find_title_row(ws, "ЯНВАРЬ 2027")
    ws.cell(row=top + 2, column=cfg.COL_STATUS).value = cfg.STATUS_READY


def break_moon_same_type(wb, ws):
    # Красим 20 сентября тем же тёмно-синим, что и ближайшее полнолуние,
    # получая два полнолуния подряд.
    top = find_title_row(ws, "СЕНТЯБРЬ 2026")
    r = top + cfg.OFFSET_DATES
    full_hex = next(h for h, v in cfg.PALETTE_DATES.items() if v == "moon_full")
    ws.cell(row=r, column=cfg.COL_FIRST_DAY + 19).fill = PatternFill(
        start_color="FF" + full_hex, end_color="FF" + full_hex, fill_type="solid")


def break_nine_activity_rows(wb, ws):
    top = find_title_row(ws, "НОЯБРЬ 2026")
    r = top + cfg.OFFSET_FIRST_ACTIVITY + CLIENT.activity_row_count
    ws.cell(row=r, column=cfg.COL_LABEL).value = "Лишняя девятая строка"


def break_seven_activity_rows(wb, ws):
    top = find_title_row(ws, "ОКТЯБРЬ 2026")
    r = top + cfg.OFFSET_FIRST_ACTIVITY + CLIENT.activity_row_count - 1
    ws.cell(row=r, column=cfg.COL_LABEL).value = None


def break_leftover_right(wb, ws):
    # Сентябрь тридцатидневный: оставляем в 31-й колонке августовскую заливку.
    top = find_title_row(ws, "СЕНТЯБРЬ 2026")
    r = top + cfg.OFFSET_FIRST_ACTIVITY
    good_hex = next(h for h, v in cfg.PALETTE_ACTIVITY.items() if v == "good")
    ws.cell(row=r, column=cfg.COL_FIRST_DAY + 30).fill = PatternFill(
        start_color="FF" + good_hex, end_color="FF" + good_hex, fill_type="solid")


def break_sheet_renamed(wb, ws):
    ws.title = "Данные за год"
    legend = wb[cfg.LEGEND_SHEET_CANDIDATES[0]]
    legend.title = "Пояснения"


def break_dates_not_number(wb, ws):
    top = find_title_row(ws, "МАЙ 2027")
    ws.cell(row=top + cfg.OFFSET_DATES, column=cfg.COL_FIRST_DAY + 4).value = "пять"


def keep_intact(wb, ws):
    """Книга остаётся целой: в этих случаях расходится конфиг, а не таблица."""


def put_client_mark(wb, ws):
    """Буква, которой нет в общем каталоге, — законна только через extraMarks."""
    top = find_title_row(ws, "ИЮЛЬ 2027")
    ws.cell(row=top + cfg.OFFSET_FIRST_ACTIVITY, column=cfg.COL_FIRST_DAY + 5).value = "Ж"


# ---------------------------------------------------------------------------
# Реестр
# ---------------------------------------------------------------------------
# expect — фрагмент, который обязан встретиться в тексте ошибки. Проверяем
# не только факт падения, но и что оно указывает на настоящую причину.
#
# Четвёртый элемент, если он есть, — конфиг клиента для этого случая. Такие
# случаи проверяют не поломку книги, а расхождение книги с конфигом: конвертер
# обязан останавливаться и на нём.

CASES = [
    ("пропущенный день",                break_missing_day,             "не подряд"),
    ("продублированный день",           break_duplicate_day,           "повторяются"),
    ("неверная длина февраля",          break_february_length,         "не високосный"),
    ("пропущенный месяц",               break_missing_month,           "не хватает"),
    ("продублированный месяц",          break_duplicate_month,         "повторяются"),
    ("неизвестный цвет",                break_unknown_color,           "не описана"),
    ("условное форматирование",         break_conditional_formatting,  "условное форматирование"),
    ("неизвестная буква",               break_unknown_mark,            "неизвестный символ"),
    ("опечатка в STATUS",               break_status_typo,             "ожидается"),
    ("два STATUS в блоке",              break_double_status,           "несколько статусов"),
    ("две одинаковые фазы подряд",      break_moon_same_type,          "чередоваться"),
    ("девять строк активностей",        break_nine_activity_rows,      "нашлась ещё одна"),
    ("семь строк активностей",          break_seven_activity_rows,     "подписи строк"),
    ("остатки справа от месяца",        break_leftover_right,          "справа от последнего дня"),
    ("переименованный лист",            break_sheet_renamed,           "не найден лист"),
    ("нечисло в строке дат",            break_dates_not_number,        "не число"),

    ("конфиг объявляет месяц, которого нет в книге",
     keep_intact, "набор месяцев не совпадает",
     lambda: client_with(range={"start": "2026-08-01", "end": "2027-09-30"})),
    ("конфиг объявляет строк больше, чем в книге",
     keep_intact, "подписи строк активностей",
     lambda: client_with(rows=ROWS + [{"id": "extra", "label": "Лишняя строка",
                                       "short": "Лишняя"}])),
    ("конфиг объявляет строк меньше, чем в книге",
     keep_intact, "нашлась ещё одна",
     lambda: client_with(rows=ROWS[:-1])),

    # Опоры луны снаружи диапазона: закрывают края, но обязаны сходиться
    # с таблицей, иначе диск на краю рисовался бы по выдуманной опоре.
    ("опора луны не чередуется с ближайшей фазой таблицы",
     keep_intact, "чередоваться",
     lambda: client_with(moonAnchors={
         "before": {"date": "2026-07-29", "type": "new"}})),
    ("опора луны слишком далеко от ближайшей фазы таблицы",
     keep_intact, "отстоит от ближайшей фазы",
     lambda: client_with(moonAnchors={
         "before": {"date": "2026-06-15", "type": "full"}})),
]

# Обратная сторона той же проверки: буква, объявленная клиентом в extraMarks,
# обязана ПРОХОДИТЬ. Иначе конфиг умел бы только запрещать.
ACCEPT_CASES = [
    ("буква из extraMarks принимается",
     put_client_mark,
     lambda: client_with(extraMarks={"Ж": "Тестовая буква этого клиента."})),
]


# ---------------------------------------------------------------------------
# Раскладка «вкладка на месяц»
# ---------------------------------------------------------------------------
# Настоящий файл астролога устроен именно так: лист «Легенда» и дальше по
# вкладке на каждый месяц. Внутри вкладки раскладка та же, что в блоке
# стопки, но статус стоит НАД заголовком, а не в его строке.

def tabs_double_status(wb, _ws):
    ws = tabs_ws(wb, "ЯНВАРЬ 2027")
    ws.cell(row=cfg.TABS_TITLE_ROW + 1, column=cfg.COL_STATUS).value = cfg.STATUS_READY


def tabs_status_typo(wb, _ws):
    ws = tabs_ws(wb, "ДЕКАБРЬ 2026")
    ws.cell(row=cfg.TABS_STATUS_ROW, column=cfg.COL_STATUS).value = "STATUS: REDY"


def tabs_duplicate_month(wb, _ws):
    # Май и март оба по 31 дню: иначе первой сработает проверка числа дней.
    tabs_ws(wb, "МАЙ 2027").cell(
        row=cfg.TABS_TITLE_ROW, column=cfg.COL_LABEL).value = "МАРТ 2027"


def tabs_missing_day(wb, _ws):
    ws = tabs_ws(wb, "СЕНТЯБРЬ 2026")
    ws.cell(row=cfg.TABS_TITLE_ROW + cfg.OFFSET_DATES,
            column=cfg.COL_FIRST_DAY + 14).value = None


def tabs_leftover_right(wb, _ws):
    # Сентябрь тридцатидневный: оставляем в 31-й колонке августовскую заливку.
    ws = tabs_ws(wb, "СЕНТЯБРЬ 2026")
    good_hex = next(h for h, v in cfg.PALETTE_ACTIVITY.items() if v == "good")
    ws.cell(row=cfg.TABS_TITLE_ROW + cfg.OFFSET_FIRST_ACTIVITY,
            column=cfg.COL_FIRST_DAY + 30).fill = PatternFill(
        start_color="FF" + good_hex, end_color="FF" + good_hex, fill_type="solid")


def tabs_conditional_formatting(wb, _ws):
    """Правило на последней вкладке, а не на первой.

    Проверка условного форматирования обязана обойти каждую вкладку: если
    она смотрит только на первую, правило на июньской пройдёт молча и месяц
    выйдет обесцвеченным.
    """
    ws = tabs_ws(wb, "ИЮНЬ 2027")
    r1 = cfg.TABS_TITLE_ROW + cfg.OFFSET_FIRST_ACTIVITY
    r2 = r1 + CLIENT.activity_row_count - 1
    ws.conditional_formatting.add(
        f"C{r1}:AG{r2}",
        CellIsRule(operator="equal", formula=['"О"'],
                   fill=PatternFill(start_color="FF00FF00", end_color="FF00FF00",
                                    fill_type="solid")))


TAB_CASES = [
    ("вкладки: два STATUS на вкладке",   tabs_double_status,          "несколько статусов"),
    ("вкладки: опечатка в STATUS",       tabs_status_typo,            "ожидается"),
    ("вкладки: месяц продублирован",     tabs_duplicate_month,        "повторяются"),
    ("вкладки: пропущенный день",        tabs_missing_day,            "не подряд"),
    ("вкладки: остатки справа",          tabs_leftover_right,         "справа от последнего дня"),
    ("вкладки: условное форматирование", tabs_conditional_formatting, "условное форматирование"),
]


def check_tabs_match_stacked():
    """Раскладка книги не должна влиять на данные — вообще ни на байт.

    Обе книги собраны из одних и тех же чисел, цветов и заметок. Если
    конвертер читает вкладки правильно, dataHash обязан совпасть: он
    считается по канонической сериализации всего содержимого.
    """
    stacked = convert(BASE, WORK / "layout-stacked.json", CLIENT)
    tabs = convert(TABS, WORK / "layout-tabs.json", CLIENT)

    assert tabs["dataHash"] == stacked["dataHash"], (
        f"вкладки дали другой dataHash:\n  стопка {stacked['dataHash']}\n"
        f"  вкладки {tabs['dataHash']}")
    return len(tabs["days"])


def check_tabs_month_without_status():
    """Вкладка без STATUS — не ошибка, а честный «месяц ещё не заполнен».

    Астролог сдаёт календарь по частям, и незаполненная вкладка обязана
    доезжать до приложения как плейсхолдер, а не ронять конвертацию.
    """
    path = fresh_copy("tabs-no-status", TABS)
    wb = load_workbook(path)
    tabs_ws(wb, "ИЮЛЬ 2027").cell(
        row=cfg.TABS_STATUS_ROW, column=cfg.COL_STATUS).value = None
    wb.save(path)

    payload = convert(path, WORK / "tabs-no-status.json", CLIENT)
    july = next(m for m in payload["months"] if m["key"] == "2027-07")
    assert july["ready"] is False, july
    assert july["note"] == "", "у незаполненного месяца заметка не выдаётся"
    assert all(m["ready"] for m in payload["months"] if m["key"] != "2027-07")


def check_dates_stored_as_floats():
    """Числа месяца, пришедшие как 1.0, — обычное дело для выгрузки из Google.

    Наивный int("1.0") падает, и весь блок перестаёт опознаваться: заголовок
    есть, а нумерации под ним «нет». Календарь развалился бы целиком.
    """
    path = fresh_copy("tabs-float-dates", TABS)
    wb = load_workbook(path)
    ws = tabs_ws(wb, "МАРТ 2027")
    row = cfg.TABS_TITLE_ROW + cfg.OFFSET_DATES
    for day in range(1, 32):
        ws.cell(row=row, column=cfg.COL_FIRST_DAY + day - 1).value = float(day)
    wb.save(path)

    payload = convert(path, WORK / "tabs-float-dates.json", CLIENT)
    assert "2027-03-31" in payload["days"], "март с числами-float не разобрался"


def check_lowercase_h_is_a_typo():
    """Строчная «х» в таблице — описка, но её нельзя лечить сплошным upper().

    В каталоге «У» (день удачи) и «у» (зуб) — разные символы с разным
    смыслом. Сплошное приведение регистра слепило бы их, и в строке зубов
    появился бы день удачи. Поднимаем регистр только там, где строчная
    форма не занята.
    """
    path = fresh_copy("tabs-lowercase", TABS)
    wb = load_workbook(path)
    ws = tabs_ws(wb, "АПРЕЛЬ 2027")
    first = cfg.TABS_TITLE_ROW + cfg.OFFSET_FIRST_ACTIVITY
    col = cfg.COL_FIRST_DAY + 2
    ws.cell(row=first, column=col).value = "Ох"                       # → ОХ
    ws.cell(row=first + CLIENT.row_ids.index("teeth"), column=col).value = "у"
    wb.save(path)

    payload = convert(path, WORK / "tabs-lowercase.json", CLIENT)
    cells = payload["days"]["2027-04-03"]["cells"]
    assert cells["negotiations"]["m"] == "ОХ", cells["negotiations"]
    assert cells["teeth"]["m"] == "у", cells["teeth"]


def check_white_beyond_month_is_not_leftover():
    """Белая заливка справа от последнего дня — не след копии августа.

    Выгрузка Google Sheets красит белым запас колонок за таблицей: в
    настоящем файле так залита целая колонка после каждого месяца. На
    экране такая ячейка неотличима от пустой, и считать её остатком —
    значит останавливать конвертацию на ровном месте.

    Цветной остаток при этом обязан ловиться по-прежнему: это проверяют
    случаи «остатки справа» выше.
    """
    path = fresh_copy("tabs-white-beyond", TABS)
    wb = load_workbook(path)
    ws = tabs_ws(wb, "СЕНТЯБРЬ 2026")          # тридцатидневный
    white = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF",
                        fill_type="solid")
    for row in (cfg.TABS_TITLE_ROW + cfg.OFFSET_DATES,
                cfg.TABS_TITLE_ROW + cfg.OFFSET_FIRST_ACTIVITY):
        ws.cell(row=row, column=cfg.COL_FIRST_DAY + 30).fill = white
    wb.save(path)

    payload = convert(path, WORK / "tabs-white-beyond.json", CLIENT)
    assert "2026-09-30" in payload["days"], "сентябрь не разобрался"


def check_moon_anchors_close_edges():
    """Опоры снаружи диапазона обязаны закрывать края календаря.

    Без них первые одиннадцать и последние четырнадцать дней остаются без
    фазы: интерполяции не на что опереться с одной стороны.
    """
    client = client_with(moonAnchors={
        "before": {"date": "2026-07-29", "type": "full"},
        "after": {"date": "2027-09-16", "type": "full"},
    })
    payload = convert(BASE, WORK / "anchored.json", client)
    dates = [e["date"] for e in payload["moonEvents"]]

    assert dates[0] == "2026-07-29", dates[:2]
    assert dates[-1] == "2027-09-16", dates[-2:]
    assert not payload["moonGaps"], payload["moonGaps"]

    # Каждый день диапазона теперь зажат между двумя опорами.
    first, last = client.range_start, client.range_end
    assert dates[0] < first, dates[0]
    assert dates[-1] > last, dates[-1]
    return len(dates)


# ---------------------------------------------------------------------------
# Запуск
# ---------------------------------------------------------------------------

def run_case(name, mutate, expect, make_client=None, base=BASE):
    slug = f"case-{abs(hash(name)) % 100000}"
    path = edit(slug, mutate, base)
    out = WORK / f"{slug}.json"
    client = make_client() if make_client else CLIENT
    try:
        convert(path, out, client)
    except ConvertError as exc:
        text = str(exc)
        if expect.lower() in text.lower():
            first = text.splitlines()[0]
            ok(f"{name}")
            print(f"    {DIM}{first[:110]}{RESET}")
            return True
        fail(f"{name}: упал, но не по той причине")
        print(f"    ожидалось вхождение «{expect}»")
        print(f"    получено: {text.splitlines()[0][:140]}")
        return False
    except Exception as exc:  # noqa: BLE001
        fail(f"{name}: упал необработанным {type(exc).__name__}: {exc}")
        return False

    fail(f"{name}: конвертер НЕ остановился — поломка прошла молча")
    return False


def run_accept(name, mutate, make_client):
    """Случай наоборот: конвертер обязан ДОРАБОТАТЬ до конца."""
    slug = f"accept-{abs(hash(name)) % 100000}"
    path = edit(slug, mutate)
    out = WORK / f"{slug}.json"
    try:
        payload = convert(path, out, make_client())
    except Exception as exc:  # noqa: BLE001
        fail(f"{name}: конвертер остановился — {type(exc).__name__}: "
             f"{str(exc).splitlines()[0][:120]}")
        return False
    ok(f"{name}")
    print(f"    {DIM}{len(payload['days'])} дней записано{RESET}")
    return True


PATH_CASES = []


def path_case(name):
    def deco(fn):
        PATH_CASES.append((name, fn))
        return fn
    return deco


@path_case("номер клиента задаёт все три пути разом")
def _():
    xlsx, config, out = resolve_paths(ROOT, "03")
    assert xlsx == ROOT / "clients" / "03" / "calendar.xlsx", xlsx
    assert config == ROOT / "clients" / "03" / "client.json", config
    assert out == ROOT / "docs" / "03" / "data" / "calendar.json", out


@path_case("каждый путь можно перебить по отдельности")
def _():
    xlsx, config, out = resolve_paths(ROOT, "03", xlsx=".tmp/proba.xlsx")
    assert xlsx == Path(".tmp/proba.xlsx"), xlsx
    # Остальные два всё равно принадлежат клиенту 03.
    assert config == ROOT / "clients" / "03" / "client.json", config


@path_case("без номера и без явного конфига — ошибка")
def _():
    try:
        resolve_paths(ROOT, None)
    except ConvertError as exc:
        assert "номер клиента" in str(exc), exc
        return
    raise AssertionError("ожидалась ошибка про неуказанного клиента")


@path_case("тесты задают все пути явно и обходятся без номера")
def _():
    xlsx, config, out = resolve_paths(
        ROOT, None,
        xlsx="ref.xlsx", config="fixtures/client-test.json", out="build/c.json")
    assert (xlsx, config, out) == (
        Path("ref.xlsx"), Path("fixtures/client-test.json"), Path("build/c.json"))


@path_case("номер вместе с чужим конфигом — ошибка")
def _():
    try:
        resolve_paths(ROOT, "03", config="clients/01/client.json")
    except ConvertError as exc:
        assert "01" in str(exc) and "03" in str(exc), exc
        return
    raise AssertionError("ожидалась ошибка про несовпадение клиента")


def main():
    WORK.mkdir(parents=True, exist_ok=True)

    head("Готовим эталонные книги")
    build(BASE, CLIENT, mode="demo", layout="stacked")
    ok(f"{BASE.name} — стопка блоков на одном листе")
    build(TABS, CLIENT, mode="demo", layout="tabs")
    ok(f"{TABS.name} — вкладка на месяц")

    head("Контрольный прогон: целая книга должна конвертироваться")
    baseline_ok = True
    try:
        payload = convert(BASE, WORK / "baseline.json", CLIENT)
        assert len(payload["days"]) == CLIENT.expected_days
        assert len(payload["months"]) == CLIENT.expected_months
        ok(f"{len(payload['days'])} дней, {len(payload['months'])} месяцев, "
           f"{len(payload['moonEvents'])} фаз")
    except Exception as exc:  # noqa: BLE001
        fail(f"эталон не конвертируется: {exc}")
        baseline_ok = False

    head("Сломанные книги")
    results = [run_case(*case) for case in CASES]

    head("Сломанные книги с вкладкой на месяц")
    results += [run_case(*case, base=TABS) for case in TAB_CASES]

    head("Случаи, где конвертер обязан пропустить")
    results += [run_accept(*case) for case in ACCEPT_CASES]

    head("Раскладка «вкладка на месяц»")
    for label, check in (
        ("вкладки дают те же данные, что стопка блоков", check_tabs_match_stacked),
        ("вкладка без STATUS — плейсхолдер, а не ошибка", check_tabs_month_without_status),
        ("числа месяца, записанные как 1.0", check_dates_stored_as_floats),
        ("строчная «х» — описка, строчная «у» — символ", check_lowercase_h_is_a_typo),
        ("белая заливка справа от месяца — не остаток",
         check_white_beyond_month_is_not_leftover),
    ):
        try:
            detail = check()
            ok(f"{label}" + (f" ({detail} дней)" if detail else ""))
            results.append(True)
        except Exception as exc:  # noqa: BLE001
            fail(f"{label}: {type(exc).__name__}: {str(exc).splitlines()[0][:140]}")
            results.append(False)

    head("Пути клиента в CLI")
    for name, fn in PATH_CASES:
        try:
            fn()
            ok(name)
            results.append(True)
        except AssertionError as exc:
            fail(f"{name}: {exc}")
            results.append(False)

    head("Опоры луны закрывают края диапазона")
    try:
        total = check_moon_anchors_close_edges()
        ok(f"края закрыты, опор стало {total}")
        results.append(True)
    except Exception as exc:  # noqa: BLE001
        fail(f"опоры не закрыли края: {type(exc).__name__}: "
             f"{str(exc).splitlines()[0][:140]}")
        results.append(False)

    passed = sum(results)
    total = len(results)
    head("Итог")
    if baseline_ok and passed == total:
        ok(f"{BOLD}все проверки пройдены: {passed}/{total}{RESET}")
        return 0
    fail(f"пройдено {passed}/{total}" + ("" if baseline_ok else ", эталон сломан"))
    return 1


if __name__ == "__main__":
    sys.exit(main())
